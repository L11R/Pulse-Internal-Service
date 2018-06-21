from openpyxl import Workbook
from django.conf import settings
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side

class BookkepingWriter(object):
    def __init__(self, name):
        self.filename = name
        
    
    def __enter__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = self.filename
        return self
        
    def dump(self, data):
        self.write_table_header(data['table_header'].values())
    
    def write_table_header(self, row):
        self.ws.append(list(row))
        #self.wb.save('{}/{}'.format(settings.FILES_ROOT, '{}.xlsx'.format(self.filename)))
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.wb.save('{}/{}'.format(settings.FILES_ROOT,'{}.xlsx'.format(self.filename)))
    

'''class BookkeepingWriter(object):
    def __init__(self, name):
        self.filename = name

    def insert_sum_by_col(self, name, sum_column):
        self._default_ws.write(self._row+1, sum_column, name)
        start_sum = xl_rowcol_to_cell(2, sum_column)
        end_sum = xl_rowcol_to_cell(self._row-1, sum_column)
        self._default_ws.write_formula(
            self._row+2,
            sum_column,
            '=SUM({}:{})'.format(start_sum, end_sum),
            self.formats['formula']
        )

    def dump(self, data):
        self.write_top_header(**data['top_header'])
        self.write_table_header(data['table_header'].values())
        for row in data['table_data']:
            self.write_table_row(row.values())


    def write_top_header(self, row, spread):
        if spread and row:
            each_size = floor(spread/len(row))
            for i, cell in enumerate(row):
                self._default_ws.merge_range(
                    self._row,
                    i*each_size,
                    self._row,
                    ((i+1)*each_size)-1,
                    cell,
                    self.formats['top_header']
                )
            self._row += 1
        else:
            if row:
                self.writerow(row, self.formats['top_header'])

    def write_table_header(self, row):
        base_opts = {
            'text_wrap': 1,
            'align': 'left',
            'valign': 'vcenter',
            'left': 1,
            'right': 1,
            'top': 2,
            'bottom': 2,
        }
        default_format = self._wb.add_format(base_opts)

        for col, cell in enumerate(row):
            newformat = None
            if col == 0:
                newformat = self._wb.add_format(base_opts)
                newformat.set_left(2)
            elif col == len(row) - 1:
                newformat = self._wb.add_format(base_opts)
                newformat.set_right(2)

            self._default_ws.set_column(col, col, min(15, len(str(cell))))
            self._default_ws.write(self._row, col, cell, newformat or default_format)

        max_len = max(sorted([len(str(x)) for x in row]))
        self._default_ws.set_row(self._row, (15*max_len//15) + 1)
        self._row += 1

    def write_table_row(self, row):
        self.writerow(row, self.formats['table_row'])

    def writerow(self, row, fmt):
        # TODO: this is way too ugly and must be rewritten.
        for col, item in enumerate(row):
            c_fmt = fmt
            if isinstance(item, (datetime, date)):
                c_fmt = copy_format(self._wb, c_fmt)
                c_fmt.set_num_format('dd.mm.yyyy')
                if item.weekday() in (5,6):
                    c_fmt.set_font_color('red')
            self._default_ws.write(self._row, col, item, c_fmt)
        self._row += 1

    def __enter__(self):
        self._wb = xlsxwriter.Workbook(self.filename,
                                       {'default_date_format': 'dd.mm.yyyy'}
        )
        self._default_ws = self._wb.add_worksheet()
        self._row = 0
        self._col = 0
        self.formats = {
            'top_header': self._wb.add_format({
                'align': 'center',
                'valign': 'vcenter',
                'bold': 1}),
            'formula': self._wb.add_format({
                'align': 'left',
                'valign': 'vcenter'}),
            'table_row': self._wb.add_format({
                'align': 'left',
                'valign': 'vcenter',
                'border': 1})
        }
        return self

    def __exit__(self, _type, value, tb):
        self._wb.close()
'''